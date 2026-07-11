"""Parse Excel files for exchange rates, IPC, liquidity, GDP, oil revenue, parallel rate."""
import io
import math
import re
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd


def _is_nan(v) -> bool:
    try:
        return math.isnan(float(v))
    except (TypeError, ValueError):
        return v is None or str(v).strip() in ("", "nan", "None", "NaN")


def _clean(v) -> Any:
    if _is_nan(v):
        return None
    return v


def _safe_float(v) -> float | None:
    if _is_nan(v):
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


MONTH_MAP_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "jun": 6, "jul": 7, "ago": 8,
    "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def detect_xlsx_type(wb: openpyxl.Workbook) -> str:
    """Detect what kind of Excel file this is by inspecting sheet names and headers."""
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Historial_TCBinance pattern
    if any("pronostico" in s or "binance" in s or "dashboard" in s for s in sheet_names_lower):
        return "rates"

    # Oil revenue + parallel rate
    if any("oil" in s or "parallel" in s or "paralelo" in s for s in sheet_names_lower):
        return "macro_oil_parallel"

    ws = wb.active
    if ws is None:
        return "unknown"

    # Read first few rows to inspect headers and content
    headers = []
    for row_idx in range(1, 5):
        row_vals = []
        for c in range(1, min(15, ws.max_column + 1)):
            cell_val = ws.cell(row_idx, c).value
            row_vals.append(str(cell_val or "").strip().lower())
        headers.extend(row_vals)

    header_str = " ".join(headers)

    if "ipc" in header_str or ("indice" in header_str and "precio" in header_str):
        return "macro_ipc"
    if "liquidez" in header_str or ("m2" in header_str and "monetari" in header_str):
        return "macro_liquidity"
    if "pib" in header_str or ("gdp" in header_str and "crecimiento" in header_str):
        return "macro_gdp"
    if "petr" in header_str or "oil" in header_str or "crudo" in header_str:
        return "macro_oil"

    # Check sheet count - VZDEC2025 has many sheets
    if len(wb.sheetnames) > 10:
        return "macro_vzdec"

    return "unknown"


def parse_rates_xlsx(content: bytes) -> dict[str, Any]:
    """Parse Historial_TCBinance Excel. Returns rows for exchange_rates table."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # Find the main data sheet (Sheet1 or first sheet with Fecha column)
    target_ws = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, min(15, ws.max_column + 1))]
        if any("fecha" in h.lower() for h in headers) and any("binance" in h.lower() or "tasa" in h.lower() for h in headers):
            target_ws = ws
            break

    if target_ws is None:
        target_ws = wb.active

    # Read headers from row 1
    headers = []
    for c in range(1, target_ws.max_column + 1):
        h = str(target_ws.cell(1, c).value or "").strip()
        headers.append(h)

    # Map column names
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if "fecha" in h_lower:
            col_map["fecha"] = i
        elif "binance" in h_lower and ("tasa" in h_lower or i == 1):
            col_map["tasa_binance"] = i
        elif ("oficial" in h_lower or "bcv" in h_lower) and ("tasa" in h_lower or i == 2):
            col_map["tasa_bcv"] = i
        elif "binance" in h_lower and ("%" in h or "dif" in h_lower):
            col_map["dif_pct_paralelo"] = i
        elif ("oficial" in h_lower or "bcv" in h_lower) and ("%" in h or "dif" in h_lower):
            col_map["dif_pct_oficial"] = i

    # Fallback
    if "tasa_binance" not in col_map:
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if "binance" in h_lower:
                col_map["tasa_binance"] = i
            elif "bcv" in h_lower or "oficial" in h_lower:
                col_map["tasa_bcv"] = i

    rows = []
    seen_dates = set()
    warnings = []

    for row_num in range(2, target_ws.max_row + 1):
        fecha_val = target_ws.cell(row_num, (col_map.get("fecha", 0)) + 1).value

        if fecha_val is None:
            continue

        fecha_str = _parse_date_value(fecha_val)
        if not fecha_str or fecha_str in seen_dates:
            continue

        seen_dates.add(fecha_str)

        def get_val(col_key):
            if col_key not in col_map:
                return None
            cell_val = target_ws.cell(row_num, col_map[col_key] + 1).value
            return _safe_float(cell_val)

        tasa_binance = get_val("tasa_binance")
        tasa_bcv = get_val("tasa_bcv")
        dif_pct_paralelo = get_val("dif_pct_paralelo")
        dif_pct_oficial = get_val("dif_pct_oficial")

        if tasa_binance is None and tasa_bcv is None:
            continue

        rows.append({
            "fecha": fecha_str,
            "tasa_binance": tasa_binance,
            "tasa_bcv": tasa_bcv,
            "dif_pct_paralelo": dif_pct_paralelo,
            "dif_pct_oficial": dif_pct_oficial,
            "log_return_binance": None,
            "log_return_bcv": None,
        })

    # Compute log returns
    for i in range(1, len(rows)):
        if rows[i]["tasa_binance"] and rows[i-1]["tasa_binance"] and rows[i-1]["tasa_binance"] > 0:
            try:
                rows[i]["log_return_binance"] = math.log(rows[i]["tasa_binance"] / rows[i-1]["tasa_binance"])
            except Exception:
                pass
        if rows[i]["tasa_bcv"] and rows[i-1]["tasa_bcv"] and rows[i-1]["tasa_bcv"] > 0:
            try:
                rows[i]["log_return_bcv"] = math.log(rows[i]["tasa_bcv"] / rows[i-1]["tasa_bcv"])
            except Exception:
                pass

    return {
        "file_type": "rates",
        "rows": rows,
        "warnings": warnings,
        "row_count": len(rows),
        "date_range": {"start": rows[0]["fecha"], "end": rows[-1]["fecha"]} if rows else None,
        "columns": list(col_map.keys()),
    }


def _parse_date_value(val) -> str | None:
    """Parse various date formats to YYYY-MM-DD string."""
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def parse_ipc_xlsx(content: bytes) -> dict[str, Any]:
    """Parse IPC.xlsx from BCV."""
    warnings = []
    rows = []

    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, header=0)
    except Exception as e:
        return {"file_type": "macro_ipc", "rows": [], "warnings": [str(e)], "row_count": 0, "columns": []}

    df.columns = [str(c).strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}

    year_col = next((col_lower[k] for k in col_lower if "año" in k or "year" in k or "anio" in k), None)
    month_col = next((col_lower[k] for k in col_lower if "mes" in k or "month" in k), None)
    index_col = next((col_lower[k] for k in col_lower if "indice" in k or "index" in k or k == "ipc"), None)
    var_col = next((col_lower[k] for k in col_lower if "var" in k and ("%" in k or "pct" in k or "porcent" in k)), None)

    for _, row in df.iterrows():
        try:
            year = int(row[year_col]) if year_col and not _is_nan(row[year_col]) else None
            month = int(row[month_col]) if month_col and not _is_nan(row[month_col]) else None
            if not year or not month or year < 2000 or month < 1 or month > 12:
                continue
            indice = _safe_float(row[index_col]) if index_col else None
            var_pct = _safe_float(row[var_col]) if var_col else None
            rows.append({"year": year, "month": month, "indice": indice, "var_pct": var_pct})
        except Exception as e:
            warnings.append(str(e)[:100])

    return {
        "file_type": "macro_ipc",
        "rows": rows,
        "warnings": warnings,
        "row_count": len(rows),
        "columns": list(df.columns),
    }


def parse_liquidity_xlsx(content: bytes) -> dict[str, Any]:
    """Parse liquidez_monetaria_mensual.xlsx (multi-sheet semester format)."""
    rows = []
    warnings = []

    try:
        xl = pd.ExcelFile(io.BytesIO(content))
    except Exception as e:
        return {"file_type": "macro_liquidity", "rows": [], "warnings": [str(e)], "row_count": 0, "columns": []}

    seen = set()

    for sheet_name in xl.sheet_names:
        try:
            df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=None)
        except Exception:
            continue

        # BCV liquidity sheets typically have date info in first column
        # and M1/M2 components in subsequent columns
        # Format varies but dates are often in col 0 as "Año/Mes" or numeric
        for row_idx in range(len(df)):
            row = df.iloc[row_idx]
            first_val = str(row.iloc[0] if len(row) > 0 else "").strip()

            year, month = _extract_year_month_from_cell(first_val, sheet_name)
            if not year or not month:
                continue

            key = (year, month)
            if key in seen:
                continue
            seen.add(key)

            # Try to extract M1/M2 values
            numeric_vals = []
            for col_idx in range(1, min(len(row), 10)):
                v = _safe_float(row.iloc[col_idx])
                if v is not None and v > 0:
                    numeric_vals.append(v)

            if not numeric_vals:
                continue

            # BCV format usually: M1, M2 components
            m1 = numeric_vals[0] if len(numeric_vals) >= 1 else None
            m2 = numeric_vals[-1] if len(numeric_vals) >= 2 else m1

            rows.append({
                "year": year, "month": month,
                "m1": m1, "m2": m2,
                "billetes_monedas": numeric_vals[0] if len(numeric_vals) >= 1 else None,
                "depositos_vista": numeric_vals[1] if len(numeric_vals) >= 2 else None,
                "depositos_ahorro": numeric_vals[2] if len(numeric_vals) >= 3 else None,
            })

    rows.sort(key=lambda r: (r["year"], r["month"]))

    return {
        "file_type": "macro_liquidity",
        "rows": rows,
        "warnings": warnings,
        "row_count": len(rows),
        "columns": ["year", "month", "m1", "m2", "billetes_monedas", "depositos_vista", "depositos_ahorro"],
    }


def _extract_year_month_from_cell(val: str, sheet_hint: str = "") -> tuple[int | None, int | None]:
    """Try to extract year and month from a cell value."""
    val = val.strip().lower()
    if not val:
        return None, None

    # Format: "2024/01" or "2024-01"
    m = re.match(r"(\d{4})[/\-](\d{1,2})$", val)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 2000 <= y <= 2030 and 1 <= mo <= 12:
            return y, mo

    # Format: "ene-24" or "enero 2024"
    for month_name, month_num in MONTH_MAP_ES.items():
        if month_name in val:
            year_match = re.search(r"(\d{4}|\d{2})", val)
            if year_match:
                yr = int(year_match.group(1))
                if yr < 100:
                    yr += 2000
                return yr, month_num

    # Extract year hint from sheet name (e.g., "I Sem 2023")
    sheet_year = None
    year_in_sheet = re.search(r"\b(20\d{2})\b", sheet_hint)
    if year_in_sheet:
        sheet_year = int(year_in_sheet.group(1))

    # If val is just a month number and sheet has year
    if sheet_year and re.match(r"^\d{1,2}$", val):
        mo = int(val)
        if 1 <= mo <= 12:
            return sheet_year, mo

    return None, None


def parse_gdp_xlsx(content: bytes) -> dict[str, Any]:
    """Parse GDP.xlsx. Returns quarterly GDP data."""
    rows = []
    warnings = []

    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, header=0)
    except Exception as e:
        return {"file_type": "macro_gdp", "rows": [], "warnings": [str(e)], "row_count": 0, "columns": []}

    df.columns = [str(c).strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}

    year_col = next((col_lower[k] for k in col_lower if "año" in k or "year" in k), None)
    quarter_col = next((col_lower[k] for k in col_lower if "trim" in k or "quarter" in k or "q" == k), None)
    gdp_col = next((col_lower[k] for k in col_lower if "pib" in k or "gdp" in k or "producto" in k), None)
    change_col = next((col_lower[k] for k in col_lower if "var" in k or "change" in k or "crec" in k), None)

    for _, row in df.iterrows():
        try:
            year = int(row[year_col]) if year_col and not _is_nan(row[year_col]) else None
            quarter = int(row[quarter_col]) if quarter_col and not _is_nan(row[quarter_col]) else None
            if not year or year < 1990:
                continue
            if not quarter:
                quarter = 1

            gdp = _safe_float(row[gdp_col]) if gdp_col else None
            pct_change = _safe_float(row[change_col]) if change_col else None

            rows.append({"year": year, "quarter": quarter, "gdp_value": gdp, "pct_change": pct_change})
        except Exception as e:
            warnings.append(str(e)[:100])

    return {
        "file_type": "macro_gdp",
        "rows": rows,
        "warnings": warnings,
        "row_count": len(rows),
        "columns": list(df.columns),
    }


def parse_oil_parallel_xlsx(content: bytes) -> dict[str, Any]:
    """Parse venezuela_oil_revenue_parallel_fx_reconstructed.xlsx."""
    oil_rows = []
    parallel_rows = []
    warnings = []

    try:
        xl = pd.ExcelFile(io.BytesIO(content))
    except Exception as e:
        return {
            "file_type": "macro_oil_parallel",
            "oil_rows": [], "parallel_rows": [],
            "warnings": [str(e)], "row_count": 0
        }

    for sheet_name in xl.sheet_names:
        sheet_lower = sheet_name.lower()
        try:
            df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=0)
            df.columns = [str(c).strip() for c in df.columns]
        except Exception:
            continue

        if "oil" in sheet_lower or "petroleo" in sheet_lower:
            # Oil Revenue sheet: Year, Revenue (USD bn), Status
            for _, row in df.iterrows():
                try:
                    year_val = row.iloc[0] if len(row) > 0 else None
                    rev_val = row.iloc[1] if len(row) > 1 else None
                    status_val = str(row.iloc[2]).strip() if len(row) > 2 else "actual"

                    year = int(_safe_float(year_val) or 0)
                    if not (2000 <= year <= 2030):
                        continue
                    revenue = _safe_float(rev_val)
                    oil_rows.append({
                        "year": year, "revenue_usd_bn": revenue, "status": status_val
                    })
                except Exception:
                    pass

        elif "parallel" in sheet_lower or "paralelo" in sheet_lower:
            # Parallel Rate Monthly sheet: Year, Month, Rate
            for _, row in df.iterrows():
                try:
                    year_val = row.iloc[0] if len(row) > 0 else None
                    month_val = row.iloc[1] if len(row) > 1 else None
                    rate_val = row.iloc[2] if len(row) > 2 else None

                    year = int(_safe_float(year_val) or 0)
                    month = int(_safe_float(month_val) or 0)
                    if not (2000 <= year <= 2030) or not (1 <= month <= 12):
                        continue
                    rate = _safe_float(rate_val)
                    parallel_rows.append({
                        "year": year, "month": month, "rate": rate, "series_type": "reconstructed"
                    })
                except Exception:
                    pass

    return {
        "file_type": "macro_oil_parallel",
        "oil_rows": oil_rows,
        "parallel_rows": parallel_rows,
        "warnings": warnings,
        "row_count": len(oil_rows) + len(parallel_rows),
    }


def import_rates(conn, rows: list[dict]) -> dict[str, int]:
    """Import exchange rate rows. Uses UPSERT (replace on conflict)."""
    imported = 0
    skipped = 0
    for row in rows:
        try:
            conn.execute(
                """INSERT OR REPLACE INTO exchange_rates
                   (fecha, tasa_binance, tasa_bcv, dif_pct_paralelo, dif_pct_oficial,
                    log_return_binance, log_return_bcv)
                   VALUES (?,?,?,?,?,?,?)""",
                (row["fecha"], row["tasa_binance"], row["tasa_bcv"],
                 row["dif_pct_paralelo"], row["dif_pct_oficial"],
                 row["log_return_binance"], row["log_return_bcv"]),
            )
            imported += 1
        except Exception:
            skipped += 1
    return {"imported": imported, "skipped": skipped}


def import_ipc(conn, rows: list[dict]) -> dict[str, int]:
    imported = 0
    for row in rows:
        conn.execute(
            "INSERT OR REPLACE INTO macro_ipc (year, month, indice, var_pct) VALUES (?,?,?,?)",
            (row["year"], row["month"], row["indice"], row["var_pct"]),
        )
        imported += 1
    return {"imported": imported, "skipped": 0}


def import_liquidity(conn, rows: list[dict]) -> dict[str, int]:
    imported = 0
    for row in rows:
        conn.execute(
            """INSERT OR REPLACE INTO macro_liquidity
               (year, month, m1, m2, billetes_monedas, depositos_vista, depositos_ahorro)
               VALUES (?,?,?,?,?,?,?)""",
            (row["year"], row["month"], row.get("m1"), row.get("m2"),
             row.get("billetes_monedas"), row.get("depositos_vista"), row.get("depositos_ahorro")),
        )
        imported += 1
    return {"imported": imported, "skipped": 0}


def import_gdp(conn, rows: list[dict]) -> dict[str, int]:
    imported = 0
    for row in rows:
        conn.execute(
            "INSERT OR REPLACE INTO macro_gdp (year, quarter, gdp_value, pct_change) VALUES (?,?,?,?)",
            (row["year"], row["quarter"], row.get("gdp_value"), row.get("pct_change")),
        )
        imported += 1
    return {"imported": imported, "skipped": 0}


def import_oil_parallel(conn, oil_rows: list[dict], parallel_rows: list[dict]) -> dict[str, int]:
    imported = 0
    for row in oil_rows:
        conn.execute(
            "INSERT OR REPLACE INTO macro_oil (year, revenue_usd_bn, status) VALUES (?,?,?)",
            (row["year"], row.get("revenue_usd_bn"), row.get("status", "actual")),
        )
        imported += 1
    for row in parallel_rows:
        conn.execute(
            "INSERT OR REPLACE INTO parallel_rate_monthly (year, month, rate, series_type) VALUES (?,?,?,?)",
            (row["year"], row["month"], row.get("rate"), row.get("series_type", "reconstructed")),
        )
        imported += 1
    return {"imported": imported, "skipped": 0}
