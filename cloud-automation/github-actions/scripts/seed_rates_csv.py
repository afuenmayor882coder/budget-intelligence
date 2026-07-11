"""Seed rates.csv from complete local history (Excel or SQLite fallback)."""
from __future__ import annotations

import argparse
import csv
import glob
import io
import math
import os
import sqlite3
import sys
from datetime import datetime, timedelta, timezone
from typing import Any

SCRIPTS_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(SCRIPTS_DIR, "..", "data")
RATES_CSV = os.path.join(DATA_DIR, "rates.csv")
DEFAULT_DB = os.path.join(SCRIPTS_DIR, "..", "..", "..", "backend", "data", "budget.db")

FIELDNAMES = [
    "fecha",
    "tasa_binance",
    "tasa_bcv",
    "dif_pct_paralelo",
    "dif_pct_oficial",
    "log_return_binance",
    "log_return_bcv",
    "updated_at",
]


def _parse_date_value(val) -> str | None:
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    if isinstance(val, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def _safe_float(val) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _compute_difs(tasa_binance: float | None, tasa_bcv: float | None) -> tuple[float | None, float | None]:
    if tasa_binance is None or tasa_bcv is None or tasa_bcv == 0 or tasa_binance == 0:
        return None, None
    return (
        round((tasa_binance - tasa_bcv) / tasa_bcv * 100, 2),
        round((tasa_bcv - tasa_binance) / tasa_binance * 100, 2),
    )


def _compute_log_return(current: float | None, previous: float | None) -> float | None:
    try:
        if current and previous and float(previous) > 0:
            return round(math.log(float(current) / float(previous)), 6)
    except (TypeError, ValueError, ZeroDivisionError):
        pass
    return None


def _enrich_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows.sort(key=lambda r: r["fecha"])
    prev_binance = None
    prev_bcv = None
    enriched = []
    now = datetime.now(timezone.utc).isoformat()

    for row in rows:
        tasa_binance = row.get("tasa_binance")
        tasa_bcv = row.get("tasa_bcv")
        dif_pct_paralelo = row.get("dif_pct_paralelo")
        dif_pct_oficial = row.get("dif_pct_oficial")

        if dif_pct_paralelo is None or dif_pct_oficial is None:
            computed_paralelo, computed_oficial = _compute_difs(tasa_binance, tasa_bcv)
            dif_pct_paralelo = dif_pct_paralelo if dif_pct_paralelo is not None else computed_paralelo
            dif_pct_oficial = dif_pct_oficial if dif_pct_oficial is not None else computed_oficial

        enriched.append({
            "fecha": row["fecha"],
            "tasa_binance": tasa_binance,
            "tasa_bcv": tasa_bcv,
            "dif_pct_paralelo": dif_pct_paralelo,
            "dif_pct_oficial": dif_pct_oficial,
            "log_return_binance": _compute_log_return(tasa_binance, prev_binance),
            "log_return_bcv": _compute_log_return(tasa_bcv, prev_bcv),
            "updated_at": row.get("updated_at") or now,
        })
        prev_binance = tasa_binance
        prev_bcv = tasa_bcv

    return enriched


def _find_excel_path(explicit: str | None) -> str | None:
    if explicit and os.path.isfile(explicit):
        return explicit

    candidates = []
    workspace_root = os.path.abspath(os.path.join(SCRIPTS_DIR, "..", "..", "..", ".."))
    patterns = [
        os.path.join(workspace_root, "Historial_TCBinance*.xlsx"),
        os.path.join(workspace_root, "Historial_TCBinance.xlsx"),
    ]
    for pattern in patterns:
        candidates.extend(glob.glob(pattern))

    if not candidates:
        return None

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def load_from_excel(path: str) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    target_ws = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, min(15, ws.max_column + 1))]
        if any("fecha" in h.lower() for h in headers) and any(
            "binance" in h.lower() or "tasa" in h.lower() for h in headers
        ):
            target_ws = ws
            break

    if target_ws is None:
        target_ws = wb.active

    headers = [str(target_ws.cell(1, c).value or "").strip() for c in range(1, target_ws.max_column + 1)]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if "fecha" in h_lower:
            col_map["fecha"] = i
        elif "binance" in h_lower and ("tasa" in h_lower or i == 1):
            col_map["tasa_binance"] = i
        elif ("oficial" in h_lower or "bcv" in h_lower) and ("tasa" in h_lower or i == 2):
            col_map["tasa_bcv"] = i
        elif "binance" in h_lower and ("%" in h or "dif" in h_lower):
            col_map["dif_pct_paralelo"] = i
        elif ("oficial" in h_lower or "bcv" in h_lower) and ("%" in h or "dif" in h_lower):
            col_map["dif_pct_oficial"] = i

    if "tasa_binance" not in col_map:
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if "binance" in h_lower:
                col_map["tasa_binance"] = i
            elif "bcv" in h_lower or "oficial" in h_lower:
                col_map["tasa_bcv"] = i

    rows: list[dict[str, Any]] = []
    seen_dates: set[str] = set()

    for row_num in range(2, target_ws.max_row + 1):
        fecha_idx = col_map.get("fecha", 0)
        fecha_val = target_ws.cell(row_num, fecha_idx + 1).value
        fecha_str = _parse_date_value(fecha_val)
        if not fecha_str or fecha_str in seen_dates:
            continue

        def get_val(col_key: str) -> float | None:
            if col_key not in col_map:
                return None
            return _safe_float(target_ws.cell(row_num, col_map[col_key] + 1).value)

        tasa_binance = get_val("tasa_binance")
        tasa_bcv = get_val("tasa_bcv")
        if tasa_binance is None and tasa_bcv is None:
            continue

        seen_dates.add(fecha_str)
        rows.append({
            "fecha": fecha_str,
            "tasa_binance": tasa_binance,
            "tasa_bcv": tasa_bcv,
            "dif_pct_paralelo": get_val("dif_pct_paralelo"),
            "dif_pct_oficial": get_val("dif_pct_oficial"),
        })

    return rows


def load_from_sqlite(db_path: str) -> list[dict[str, Any]]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.execute(
        """SELECT fecha, tasa_binance, tasa_bcv, dif_pct_paralelo, dif_pct_oficial
           FROM exchange_rates ORDER BY fecha"""
    )
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows


def validate_rows(rows: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    if not rows:
        warnings.append("No rows to export")
        return warnings

    fechas = [r["fecha"] for r in rows]
    if len(fechas) != len(set(fechas)):
        warnings.append(f"Duplicate dates found ({len(fechas) - len(set(fechas))} duplicates)")

    rows_sorted = sorted(rows, key=lambda r: r["fecha"])
    start = datetime.strptime(rows_sorted[0]["fecha"], "%Y-%m-%d").date()
    end = datetime.strptime(rows_sorted[-1]["fecha"], "%Y-%m-%d").date()
    expected_days = (end - start).days + 1
    if len(rows_sorted) != expected_days:
        missing = expected_days - len(rows_sorted)
        warnings.append(
            f"Date coverage gap: {len(rows_sorted)} rows from {start} to {end}, "
            f"expected {expected_days} calendar days ({missing} missing)"
        )

    return warnings


def write_csv(rows: list[dict[str, Any]], output_path: str = RATES_CSV) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k) for k in FIELDNAMES})


def main() -> int:
    parser = argparse.ArgumentParser(description="Seed rates.csv from Excel or SQLite")
    parser.add_argument("--excel", help="Path to Historial_TCBinance.xlsx")
    parser.add_argument("--db", default=DEFAULT_DB, help="SQLite fallback path")
    parser.add_argument("--output", default=RATES_CSV, help="Output CSV path")
    args = parser.parse_args()

    source = "unknown"
    raw_rows: list[dict[str, Any]] = []

    excel_path = _find_excel_path(args.excel)
    if excel_path:
        print(f"Reading Excel: {excel_path}")
        raw_rows = load_from_excel(excel_path)
        source = f"excel:{excel_path}"
    elif os.path.isfile(args.db):
        print(f"Excel not found; falling back to SQLite: {args.db}")
        raw_rows = load_from_sqlite(args.db)
        source = f"sqlite:{args.db}"
    else:
        print("ERROR: No Excel file found and SQLite DB missing", file=sys.stderr)
        return 1

    if not raw_rows:
        print("ERROR: No rate rows parsed from source", file=sys.stderr)
        return 1

    rows = _enrich_rows(raw_rows)
    warnings = validate_rows(rows)
    for w in warnings:
        print(f"WARNING: {w}", file=sys.stderr)

    write_csv(rows, args.output)
    print(
        f"Seeded {args.output}: {len(rows)} rows "
        f"({rows[0]['fecha']} → {rows[-1]['fecha']}) from {source}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
